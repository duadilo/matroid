"""
Matroid Excel Processing Server
================================
FastAPI + uvicorn server that exposes Excel operations over HTTP.

Startup contract with the Flutter host
---------------------------------------
When the server is ready it prints exactly one line to stdout:

    PORT:<n>

where <n> is the dynamically chosen port. The Flutter PythonServer class
reads this line (with a 10-second timeout) to know which port to use.

Build
------
See build.sh for the PyInstaller command that packages this into a
self-contained binary placed alongside the Flutter executable.
"""

import asyncio
import base64
import csv
import io
import json
import os
import re
import shutil
import socket
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from typing import Any, AsyncGenerator

import openpyxl
import uvicorn
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from RestrictedPython import compile_restricted, safe_builtins, safe_globals
from RestrictedPython.Guards import full_write_guard, safer_getattr
from RestrictedPython.PrintCollector import PrintCollector

app = FastAPI(title="Matroid Excel Server", docs_url=None, redoc_url=None)

_MAX_UPLOAD_BYTES = 50 * 1024 * 1024  # 50 MB
_MAX_TIMEOUT_SECONDS = 60

# ---------------------------------------------------------------------------
# In-memory state
# ---------------------------------------------------------------------------

_workbook: dict[str, Any] = {
    "bytes": None,
    "name": None,
}


def _require_workbook() -> None:
    if _workbook["bytes"] is None:
        raise HTTPException(status_code=400, detail="No file loaded. Call /load first.")


# ---------------------------------------------------------------------------
# Workbook helper functions (injected as globals into user scripts)
# ---------------------------------------------------------------------------


def _wb_get_data() -> list[dict[str, Any]]:
    """Return all rows (after the header) as a list of dicts."""
    if _workbook["bytes"] is None:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(_workbook["bytes"]), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def _wb_get_sheet_names() -> list[str]:
    if _workbook["bytes"] is None:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(_workbook["bytes"]), read_only=True)
    return wb.sheetnames


def _wb_get_headers() -> list[str]:
    if _workbook["bytes"] is None:
        return []
    wb = openpyxl.load_workbook(io.BytesIO(_workbook["bytes"]), read_only=True, data_only=True)
    ws = wb.active
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(h) if h is not None else f"col{i}" for i, h in enumerate(first_row)]


def _wb_to_csv() -> str:
    data = _wb_get_data()
    if not data:
        return ""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(data[0].keys()))
    writer.writeheader()
    writer.writerows(data)
    return buf.getvalue()


def _wb_summarize() -> dict[str, Any]:
    data = _wb_get_data()
    if not data:
        return {}
    result: dict[str, Any] = {}
    for key in data[0].keys():
        values = [row.get(key) for row in data if row.get(key) is not None]
        result[key] = {
            "count": len(values),
            "unique": len({str(v) for v in values}),
            "sample": values[:3],
        }
    return result


def _wb_filter_rows(column: str, value: Any) -> list[dict[str, Any]]:
    return [row for row in _wb_get_data() if str(row.get(column, "")) == str(value)]


# ---------------------------------------------------------------------------
# Execution helpers
# ---------------------------------------------------------------------------


def _run_python(code: str, timeout: int) -> dict[str, Any]:
    """Execute user Python code in a RestrictedPython sandbox."""
    # Compile first — catches syntax errors before we touch a thread.
    try:
        compiled = compile_restricted(code, "<user_code>", "exec")
    except SyntaxError as exc:
        return {"stdout": "", "stderr": "", "error": f"SyntaxError: {exc}", "execution_time_ms": 0}

    # Restricted globals: safe builtins + our Excel helpers.
    glb: dict[str, Any] = {
        "__builtins__": safe_builtins,
        "_print_": PrintCollector,
        "_getiter_": iter,
        "_getattr_": safer_getattr,
        "_getitem_": lambda obj, key: obj[key],
        "_write_": full_write_guard,
        "get_data": _wb_get_data,
        "get_sheet_names": _wb_get_sheet_names,
        "get_headers": _wb_get_headers,
        "to_csv": _wb_to_csv,
        "summarize": _wb_summarize,
        "filter_rows": _wb_filter_rows,
    }

    start = time.monotonic()

    def _execute() -> str:
        exec(compiled, glb)  # noqa: S102
        # PrintCollector accumulates output; calling _print_() returns the string.
        printer = glb.get("_print")
        return printer() if callable(printer) else ""

    # Use shutdown(wait=False) so an infinite-loop thread doesn't block the
    # API response after the timeout fires.
    executor = ThreadPoolExecutor(max_workers=1)
    future = executor.submit(_execute)
    try:
        stdout = future.result(timeout=timeout)
        elapsed = int((time.monotonic() - start) * 1000)
        executor.shutdown(wait=False)
        return {"stdout": stdout, "stderr": "", "error": None, "execution_time_ms": elapsed}
    except FuturesTimeoutError:
        elapsed = int((time.monotonic() - start) * 1000)
        executor.shutdown(wait=False)
        return {
            "stdout": "",
            "stderr": "",
            "error": f"Execution timed out after {timeout}s",
            "execution_time_ms": elapsed,
        }
    except Exception as exc:  # noqa: BLE001
        elapsed = int((time.monotonic() - start) * 1000)
        executor.shutdown(wait=False)
        return {"stdout": "", "stderr": str(exc), "error": str(exc), "execution_time_ms": elapsed}


def _run_javascript(code: str, timeout: int) -> dict[str, Any]:
    """Execute user JavaScript code in a Node.js subprocess."""
    node_path = shutil.which("node")
    if node_path is None:
        return {
            "stdout": "",
            "stderr": "",
            "error": "Node.js not found. Install Node.js to run JavaScript.",
            "execution_time_ms": 0,
        }

    # Build the preamble that exposes Excel helper globals to the user's script.
    data = _wb_get_data()
    headers = _wb_get_headers()
    sheets = _wb_get_sheet_names()

    preamble = f"""\
const _data={json.dumps(data, default=str)};
const _headers={json.dumps(headers)};
const _sheets={json.dumps(sheets)};
function getData(){{return _data;}}
function getSheetNames(){{return _sheets;}}
function getHeaders(){{return _headers;}}
function toCsv(){{
  if(!_data.length)return'';
  const h=Object.keys(_data[0]);
  return[h.join(','),..._data.map(r=>h.map(k=>JSON.stringify(r[k]??'')).join(','))].join('\\n');
}}
function summarize(){{
  const r={{}};
  _headers.forEach(h=>{{
    const v=_data.map(row=>row[h]).filter(x=>x!=null);
    r[h]={{count:v.length,unique:new Set(v.map(String)).size,sample:v.slice(0,3)}};
  }});
  return r;
}}
function filterRows(col,val){{return _data.filter(r=>String(r[col]??'')===String(val));}}
"""
    full_code = preamble + "\n(async () => {\n" + code + "\n})().catch(e => { process.stderr.write(String(e)); process.exit(1); });"
    start = time.monotonic()
    try:
        result = subprocess.run(
            [node_path, "-e", full_code],
            capture_output=True,
            text=True,
            timeout=timeout,
        )
        elapsed = int((time.monotonic() - start) * 1000)
        error = result.stderr.strip() if result.returncode != 0 else None
        return {
            "stdout": result.stdout,
            "stderr": result.stderr,
            "error": error,
            "execution_time_ms": elapsed,
        }
    except subprocess.TimeoutExpired:
        elapsed = int((time.monotonic() - start) * 1000)
        return {
            "stdout": "",
            "stderr": "",
            "error": f"Execution timed out after {timeout}s",
            "execution_time_ms": elapsed,
        }


# ---------------------------------------------------------------------------
# Request / response models
# ---------------------------------------------------------------------------


class ExecuteRequest(BaseModel):
    language: str
    code: str
    timeout: int = 10


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------


@app.get("/health")
async def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/load")
async def load_file(file: UploadFile = File(...)) -> dict[str, Any]:
    data = await file.read()
    if len(data) > _MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large ({len(data)} bytes). Maximum is {_MAX_UPLOAD_BYTES} bytes.",
        )
    _workbook["bytes"] = data
    _workbook["name"] = file.filename
    return {
        "status": "loaded",
        "filename": file.filename,
        "size_bytes": len(data),
    }


@app.post("/process")
async def process(request: Request) -> dict[str, Any]:
    _require_workbook()

    try:
        params: dict[str, Any] = await request.json()
    except Exception:
        params = {}

    # ---------------------------------------------------------------------------
    # TODO: replace this stub with real openpyxl / pandas logic.
    # ---------------------------------------------------------------------------

    return {
        "status": "processed",
        "filename": _workbook["name"],
        "params": params,
        "rows": 0,  # placeholder
    }


@app.get("/export")
async def export() -> Response:
    _require_workbook()
    safe_name = re.sub(r'[^\w.\-]', '_', _workbook["name"] or "export.xlsx")
    return Response(
        content=_workbook["bytes"],
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{safe_name}"'
        },
    )


@app.post("/unload")
async def unload() -> dict[str, str]:
    _workbook["bytes"] = None
    _workbook["name"] = None
    return {"status": "unloaded"}


@app.post("/execute")
async def execute(req: ExecuteRequest) -> dict[str, Any]:
    """Run user-supplied code in a sandboxed environment.

    Python: RestrictedPython sandbox with Excel helper globals.
    JavaScript: Node.js subprocess (requires Node installed on the server).

    The response always has HTTP 200; execution errors are reported in the
    ``error`` field so the Flutter client can display them in the output panel.
    """
    timeout = min(max(req.timeout, 1), _MAX_TIMEOUT_SECONDS)
    if req.language == "python":
        return _run_python(req.code, timeout)
    if req.language == "javascript":
        return _run_javascript(req.code, timeout)
    raise HTTPException(
        status_code=400,
        detail=f"Unsupported language '{req.language}'. Use 'python' or 'javascript'.",
    )


# ---------------------------------------------------------------------------
# Chat — Pydantic models
# ---------------------------------------------------------------------------


class ChatAttachment(BaseModel):
    media_type: str
    data: str  # base64


class ChatMessage(BaseModel):
    role: str
    content: str
    attachments: list[ChatAttachment] = []


class ChatStreamRequest(BaseModel):
    provider: str  # "openai" | "anthropic" | "gemini"
    model: str
    messages: list[ChatMessage]
    base_url: str | None = None
    system_prompt: str | None = None
    api_keys: dict[str, str] | None = None  # client-side overrides
    tools_enabled: bool = False


# ---------------------------------------------------------------------------
# Chat — API key resolution
# ---------------------------------------------------------------------------

_ENV_KEY_MAP = {
    "openai": "OPENAI_API_KEY",
    "anthropic": "ANTHROPIC_API_KEY",
    "gemini": "GOOGLE_API_KEY",
    "qwen": "QWEN_API_KEY",  # Optional: set if your custom endpoint requires auth
}

_DEFAULT_MODELS: dict[str, list[str]] = {
    "openai": ["gpt-4o", "gpt-4o-mini", "gpt-4.1", "gpt-4.1-mini", "gpt-4.1-nano", "o3-mini"],
    "anthropic": [
        "claude-sonnet-4-20250514",
        "claude-haiku-4-20250414",
        "claude-opus-4-20250514",
    ],
    "gemini": ["gemini-2.0-flash", "gemini-2.5-pro-preview-06-05"],
    "qwen": ["qwen3.5-122b-a10b"],
}


def _get_api_key(provider: str, req: ChatStreamRequest) -> str | None:
    """Return the API key for *provider* — client override first, then env."""
    if req.api_keys and provider in req.api_keys:
        return req.api_keys[provider] or None
    env_var = _ENV_KEY_MAP.get(provider)
    return os.environ.get(env_var) if env_var else None


# ---------------------------------------------------------------------------
# Chat — Web search tool
# ---------------------------------------------------------------------------


def _web_search(query: str, max_results: int = 5) -> str:
    """Execute a web search using DuckDuckGo and return formatted results."""
    from duckduckgo_search import DDGS

    with DDGS() as ddgs:
        results = list(ddgs.text(query, max_results=max_results))
    if not results:
        return "No results found."
    return "\n\n".join(
        f"[{r['title']}]({r['href']})\n{r['body']}" for r in results
    )


# ---------------------------------------------------------------------------
# Chat — Provider streaming functions
# ---------------------------------------------------------------------------


async def _stream_openai(req: ChatStreamRequest) -> AsyncGenerator[tuple[str, Any], None]:
    import openai

    api_key = _get_api_key("openai", req)
    # Allow empty API key if custom base_url is provided (for local/OpenAI-compatible endpoints)
    if not api_key and not req.base_url:
        raise ValueError("OpenAI API key not configured")
    # Use a dummy key if none provided but custom endpoint is set
    effective_api_key = api_key or "dummy-key-for-custom-endpoint"
    client = openai.AsyncOpenAI(api_key=effective_api_key, base_url=req.base_url)

    messages: list[dict[str, Any]] = []
    if req.system_prompt:
        messages.append({"role": "system", "content": req.system_prompt})
    for msg in req.messages:
        if msg.attachments:
            content: list[dict[str, Any]] = [{"type": "text", "text": msg.content}]
            for att in msg.attachments:
                content.append(
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{att.media_type};base64,{att.data}"
                        },
                    }
                )
            messages.append({"role": msg.role, "content": content})
        else:
            messages.append({"role": msg.role, "content": msg.content})

    tools = None
    if req.tools_enabled:
        tools = [
            {
                "type": "function",
                "function": {
                    "name": "web_search",
                    "description": "Search the web for current information",
                    "parameters": {
                        "type": "object",
                        "properties": {"query": {"type": "string"}},
                        "required": ["query"],
                    },
                },
            }
        ]

    kwargs: dict[str, Any] = {"model": req.model, "messages": messages, "stream": True}
    if tools:
        kwargs["tools"] = tools

    while True:
        stream = await client.chat.completions.create(**kwargs)
        tool_calls_acc: dict[int, dict[str, str]] = {}  # index → {id, name, arguments}
        finish_reason = None

        async for chunk in stream:
            choice = chunk.choices[0] if chunk.choices else None
            if not choice:
                continue
            if choice.finish_reason:
                finish_reason = choice.finish_reason
            delta = choice.delta
            if delta and delta.content:
                yield ("text", delta.content)
            if delta and delta.tool_calls:
                for tc in delta.tool_calls:
                    idx = tc.index
                    if idx not in tool_calls_acc:
                        tool_calls_acc[idx] = {"id": "", "name": "", "arguments": ""}
                    if tc.id:
                        tool_calls_acc[idx]["id"] = tc.id
                    if tc.function and tc.function.name:
                        tool_calls_acc[idx]["name"] = tc.function.name
                    if tc.function and tc.function.arguments:
                        tool_calls_acc[idx]["arguments"] += tc.function.arguments

        if finish_reason != "tool_calls" or not tool_calls_acc:
            break

        # Execute tool calls and feed results back
        assistant_msg: dict[str, Any] = {
            "role": "assistant",
            "content": None,
            "tool_calls": [
                {
                    "id": tc["id"],
                    "type": "function",
                    "function": {"name": tc["name"], "arguments": tc["arguments"]},
                }
                for tc in tool_calls_acc.values()
            ],
        }
        messages.append(assistant_msg)

        for tc in tool_calls_acc.values():
            args = json.loads(tc["arguments"])
            if tc["name"] == "web_search":
                query = args.get("query", "")
                yield ("tool_use", {"name": "web_search", "input": {"query": query}})
                result = await asyncio.to_thread(_web_search, query)
                yield ("tool_result", {"name": "web_search"})
            else:
                result = f"Unknown tool: {tc['name']}"
            messages.append({
                "role": "tool",
                "tool_call_id": tc["id"],
                "content": result,
            })

        kwargs["messages"] = messages


async def _stream_anthropic(req: ChatStreamRequest) -> AsyncGenerator[tuple[str, Any], None]:
    import anthropic

    api_key = _get_api_key("anthropic", req)
    if not api_key:
        raise ValueError("Anthropic API key not configured")
    client = anthropic.AsyncAnthropic(api_key=api_key)

    messages: list[dict[str, Any]] = []
    for msg in req.messages:
        if msg.attachments:
            content: list[dict[str, Any]] = []
            for att in msg.attachments:
                content.append(
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": att.media_type,
                            "data": att.data,
                        },
                    }
                )
            content.append({"type": "text", "text": msg.content})
            messages.append({"role": msg.role, "content": content})
        else:
            messages.append({"role": msg.role, "content": msg.content})

    kwargs: dict[str, Any] = {
        "model": req.model,
        "messages": messages,
        "max_tokens": 4096,
    }
    if req.system_prompt:
        kwargs["system"] = req.system_prompt

    if req.tools_enabled:
        kwargs["tools"] = [
            {"type": "web_search_20250305", "name": "web_search", "max_uses": 5}
        ]

    if not req.tools_enabled:
        async with client.messages.stream(**kwargs) as stream:
            async for text in stream.text_stream:
                yield ("text", text)
    else:
        async with client.messages.stream(**kwargs) as stream:
            async for event in stream:
                if event.type == "content_block_start":
                    block = event.content_block
                    if hasattr(block, "type") and block.type == "server_tool_use":
                        yield ("tool_use", {"name": block.name, "input": getattr(block, "input", {})})
                elif event.type == "content_block_stop":
                    # Check if the stopped block was a tool result
                    pass
                elif event.type == "text":
                    yield ("text", event.text)
            # After the stream, check the final message for tool results
            msg = await stream.get_final_message()
            for block in msg.content:
                if hasattr(block, "type") and block.type == "web_search_tool_result":
                    yield ("tool_result", {"name": "web_search"})


async def _stream_gemini(req: ChatStreamRequest) -> AsyncGenerator[tuple[str, Any], None]:
    import google.generativeai as genai

    api_key = _get_api_key("gemini", req)
    if not api_key:
        raise ValueError("Google AI API key not configured")
    genai.configure(api_key=api_key)

    tools_kwarg = None
    if req.tools_enabled:
        tools_kwarg = genai.types.Tool(
            function_declarations=[
                genai.types.FunctionDeclaration(
                    name="web_search",
                    description="Search the web for current information",
                    parameters={
                        "type": "object",
                        "properties": {"query": {"type": "string"}},
                        "required": ["query"],
                    },
                )
            ]
        )

    model = genai.GenerativeModel(
        req.model,
        system_instruction=req.system_prompt if req.system_prompt else None,
        tools=[tools_kwarg] if tools_kwarg else None,
    )

    contents: list[dict[str, Any]] = []
    for msg in req.messages:
        role = "user" if msg.role == "user" else "model"
        parts: list[Any] = [msg.content]
        for att in msg.attachments:
            raw = base64.b64decode(att.data)
            parts.append({"mime_type": att.media_type, "data": raw})
        contents.append({"role": role, "parts": parts})

    chat = model.start_chat(history=contents[:-1] if len(contents) > 1 else [])
    last_user = contents[-1] if contents else {"role": "user", "parts": [""]}

    while True:
        response = await asyncio.to_thread(
            lambda parts=last_user["parts"]: chat.send_message(parts, stream=True)
        )
        has_function_call = False
        for chunk in response:
            for part in chunk.parts:
                if hasattr(part, "function_call") and part.function_call.name:
                    fc = part.function_call
                    query = dict(fc.args).get("query", "")
                    yield ("tool_use", {"name": "web_search", "input": {"query": query}})
                    result = _web_search(query)
                    yield ("tool_result", {"name": "web_search"})
                    last_user = {
                        "role": "user",
                        "parts": [genai.types.content_types.to_part(
                            genai.types.FunctionResponse(name="web_search", response={"result": result})
                        )],
                    }
                    has_function_call = True
                elif hasattr(part, "text") and part.text:
                    yield ("text", part.text)

        if not has_function_call:
            break


_STREAM_PROVIDERS = {
    "openai": _stream_openai,
    "anthropic": _stream_anthropic,
    "gemini": _stream_gemini,
    "qwen": _stream_openai,  # Qwen is OpenAI-compatible, reuse the same handler
}


# ---------------------------------------------------------------------------
# Chat — Endpoints
# ---------------------------------------------------------------------------


@app.get("/chat/models")
async def chat_models() -> dict[str, Any]:
    """Return available models per provider based on configured API keys."""
    providers: dict[str, list[str]] = {}
    for provider, env_var in _ENV_KEY_MAP.items():
        if provider == "qwen":
            # Always include qwen (custom endpoint doesn't require env var)
            providers[provider] = _DEFAULT_MODELS[provider]
        elif os.environ.get(env_var):
            providers[provider] = _DEFAULT_MODELS[provider]
    return {"providers": providers}


@app.post("/chat/stream")
async def chat_stream(req: ChatStreamRequest) -> StreamingResponse:
    stream_fn = _STREAM_PROVIDERS.get(req.provider)
    if stream_fn is None:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown provider '{req.provider}'. Use one of: {list(_STREAM_PROVIDERS)}",
        )

    async def generate() -> AsyncGenerator[str, None]:
        try:
            async for event_type, payload in stream_fn(req):
                if event_type == "text":
                    escaped = payload.replace("\n", "\\n")
                    yield f"data: {escaped}\n\n"
                elif event_type == "tool_use":
                    yield f"event: tool_use\ndata: {json.dumps(payload)}\n\n"
                elif event_type == "tool_result":
                    yield f"event: tool_result\ndata: {json.dumps(payload)}\n\n"
            yield "data: [DONE]\n\n"
        except Exception as exc:
            yield f"event: error\ndata: {exc!s}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def _free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("", 0))
        return s.getsockname()[1]


if __name__ == "__main__":
    port = _free_port()

    print(f"PORT:{port}", flush=True)

    uvicorn.run(
        app,
        host="127.0.0.1",
        port=port,
        log_level="error",
        access_log=False,
    )
